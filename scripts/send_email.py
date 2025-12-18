#!/usr/bin/env python3
"""
Send Excel report via email after scraping.
Uses SMTP configuration from config.json or environment variables.
"""
import smtplib
import ssl
import os
import json
import sys
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from email.utils import formataddr
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


def load_config():
    """Load SMTP config from config.json or environment variables."""
    config = {}
    config_path = Path(__file__).parent.parent / 'config.json'
    
    if config_path.exists():
        with open(config_path, 'r') as f:
            config = json.load(f)
    
    return {
        'smtp_host': config.get('smtp_host', os.environ.get('SMTP_HOST', 'ssl0.ovh.net')),
        'smtp_port': int(config.get('smtp_port', os.environ.get('SMTP_PORT', 465))),
        'smtp_user': config.get('smtp_user', os.environ.get('SMTP_USER')),
        'smtp_pass': config.get('smtp_pass', os.environ.get('SMTP_PASS')),
        'recipient': config.get('recipient', os.environ.get('RECIPIENT', 'fabien.voyer@orange.com')),
    }


def send_excel_report(excel_path: Path, row_count: int):
    """Send the Excel file as an email attachment."""
    config = load_config()
    
    if not config['smtp_user'] or not config['smtp_pass']:
        print("⚠️ SMTP credentials not configured. Email not sent.")
        print("   Configure config.json or set SMTP_USER/SMTP_PASS environment variables.")
        return False
    
    # Create message
    msg = MIMEMultipart()
    msg['Subject'] = f"📊 Telecom Market Cap Report - {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    msg['From'] = formataddr(("Market Cap Scraper", config['smtp_user']))
    msg['To'] = config['recipient']
    
    # Email body
    body = f"""Bonjour,

Le scraping des capitalisations boursières des entreprises de télécommunications est terminé.

📊 Résultats:
- {row_count} entreprises extraites
- Fichier: {excel_path.name}
- Date: {datetime.now().strftime('%d/%m/%Y à %H:%M')}

Le fichier Excel est joint à cet email.

---
Ce message est envoyé automatiquement par le script de scraping.
"""
    msg.attach(MIMEText(body, 'plain', 'utf-8'))
    
    # Attach Excel file
    if excel_path.exists():
        with open(excel_path, 'rb') as f:
            part = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header(
            'Content-Disposition',
            f'attachment; filename="{excel_path.name}"'
        )
        msg.attach(part)
    
    # Send email
    try:
        context = ssl.create_default_context()
        with smtplib.SMTP_SSL(config['smtp_host'], config['smtp_port'], context=context) as server:
            server.login(config['smtp_user'], config['smtp_pass'])
            server.sendmail(config['smtp_user'], [config['recipient']], msg.as_string())
        
        print(f"✅ Email envoyé à {config['recipient']}")
        return True
    except Exception as e:
        print(f"❌ Erreur d'envoi d'email: {e}")
        return False


if __name__ == "__main__":
    # Test standalone
    excel_file = Path(__file__).parent.parent / "telecom_market_caps_eur_billion.xlsx"
    if excel_file.exists():
        # Count rows in Excel (minus header and timestamp row)
        workbook = load_workbook(excel_file)
        worksheet = workbook.active
        row_count = worksheet.max_row - 2  # minus header and timestamp row
        send_excel_report(excel_file, row_count)
    else:
        print(f"Excel file not found: {excel_file}")
