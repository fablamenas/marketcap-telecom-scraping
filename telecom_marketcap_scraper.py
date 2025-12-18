from __future__ import annotations

import argparse
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable
from zoneinfo import ZoneInfo

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

BASE_URL = (
    "https://companiesmarketcap.com/fr/telecommunication/"
    "plus-grandes-entreprises-de-telecommunications-par-capitalisation-boursiere/"
)
TZ = ZoneInfo("Europe/Paris")
DEFAULT_TIMEOUT = 30
MAX_PAGES = 20
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36"
    )
}


@dataclass
class CompanyRow:
    rank: int
    name: str
    market_cap_billion_eur: float
    country: str


def parse_market_cap_to_billion_eur(text: str) -> float | None:
    """Return the market cap value expressed in billions of euros.

    The website sometimes labels billions as "Md" (milliards) or "Bn" (billion).
    """

    normalized = text.replace("\xa0", " ").strip()
    match = re.search(r"€\s*([0-9]+(?:[.,][0-9]+)?)\s*(Md|Bn)", normalized)
    if not match:
        return None

    value = float(match.group(1).replace(",", "."))
    unit = match.group(2)

    # Md and Bn are both billions; keep value unchanged for now in case the
    # source mixes them.
    if unit not in {"Md", "Bn"}:
        return None

    return value


def fetch_page(page: int, *, timeout: int) -> BeautifulSoup:
    url = BASE_URL if page == 1 else f"{BASE_URL}?page={page}"
    response = requests.get(url, timeout=timeout, headers=HEADERS)
    response.raise_for_status()
    return BeautifulSoup(response.text, "html.parser")


def extract_rows(soup: BeautifulSoup) -> list[CompanyRow]:
    table = soup.find("table")
    if not table:
        return []

    rows: list[CompanyRow] = []
    for tr in table.select("tbody tr"):
        tds = tr.find_all("td")
        if len(tds) < 8:
            continue

        # Column structure: 0=empty, 1=rank, 2=name+ticker, 3=market_cap, 4=price, 5=today%, 6=30days, 7=country
        rank_text = tds[1].get_text(strip=True)
        name_cell = tds[2].get_text(" ", strip=True)
        market_cap_text = tds[3].get_text(" ", strip=True)
        country_raw = tds[7].get_text(" ", strip=True)
        # Remove flag emoji (first character + space) for Excel compatibility
        country = re.sub(r'^[\U0001F1E0-\U0001F1FF]{2}\s*', '', country_raw).strip()

        market_cap = parse_market_cap_to_billion_eur(market_cap_text)
        if not rank_text or market_cap is None:
            continue

        name_without_ticker = re.sub(r"\s+[A-Z0-9.\-]+$", "", name_cell).strip()

        try:
            rank = int(rank_text)
        except ValueError:
            continue

        rows.append(
            CompanyRow(
                rank=rank,
                name=name_without_ticker,
                market_cap_billion_eur=market_cap,
                country=country,
            )
        )

    return rows


def scrape_all(*, timeout: int = DEFAULT_TIMEOUT, max_pages: int = MAX_PAGES) -> list[CompanyRow]:
    all_rows: list[CompanyRow] = []
    page = 1

    while True:
        soup = fetch_page(page, timeout=timeout)
        rows = extract_rows(soup)
        if not rows:
            break

        all_rows.extend(rows)
        page += 1

        if page > max_pages:
            break

    return all_rows


def write_excel(rows: Iterable[CompanyRow], *, output: Path, extracted_at: datetime) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)

    sorted_rows = sorted(rows, key=lambda row: row.rank)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "market_caps"

    worksheet.append(["rang", "nom", "market_cap_(Mds_EUR)", "pays"])
    worksheet.append(
        [
            "extracted_at_europe_paris",
            extracted_at.isoformat(timespec="seconds"),
            "",
            "",
        ]
    )

    for row in sorted_rows:
        worksheet.append(
            [
                row.rank,
                row.name,
                round(row.market_cap_billion_eur, 2),
                row.country,
            ]
        )
        worksheet.cell(row=worksheet.max_row, column=3).number_format = "0.00"

    workbook.save(output)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Scrape telecom market caps and write them to an Excel file."
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("telecom_market_caps_eur_billion.xlsx"),
        help="Output Excel path (default: telecom_market_caps_eur_billion.xlsx)",
    )
    parser.add_argument(
        "--max-pages",
        type=int,
        default=MAX_PAGES,
        help=f"Maximum number of pages to scrape (default: {MAX_PAGES})",
    )
    parser.add_argument(
        "--timeout",
        type=int,
        default=DEFAULT_TIMEOUT,
        help=f"HTTP timeout in seconds (default: {DEFAULT_TIMEOUT})",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    extracted_at = datetime.now(TZ)

    rows = scrape_all(timeout=args.timeout, max_pages=args.max_pages)
    write_excel(rows, output=args.output, extracted_at=extracted_at)

    print(f"{len(rows)} lignes écrites dans {args.output}")


if __name__ == "__main__":
    main()
